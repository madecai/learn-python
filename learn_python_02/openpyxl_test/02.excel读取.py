import openpyxl

# data_only=True读取数据而不是公式的计算结果
book = openpyxl.load_workbook('test01.xlsx', data_only=True)

sheet = book.active

# 获取方式一
a1 = sheet['A1']
a2 = sheet['A2']
# 获取方式二
a3 = sheet.cell(row=3, column=1)
print(a1.value, a2.value, a3.value)

# 获取方式三
print('===========区间获取，返回每行cell的元组')
cells = sheet['A1': 'B6']

for cella, cellb in cells:
    print(cella.value, cellb.value)

# 获取方式四
print('===================按行迭代,返回每行cell的元组')
for cella, cellb, cellc in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
	print(cella.value, cellb.value, cellc.value)

# 获取方式五
print('===================按列迭代,返回每列cell的元组')
for cols in sheet.iter_cols(min_row=1, min_col=1, max_row=6, max_col=3):
	for cell in cols:
		print(cell.value)
	print('----------')

# 获取方式六
print('===================获取全部行,返回每列cell的元组')
rows = sheet.rows
for row in rows:
	for cell in row:
		print(cell.value, end=" ")
	print()