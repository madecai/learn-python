import openpyxl

# 创建完就默认有一个名为“Sheet”的页
wb = openpyxl.Workbook()
wb.create_sheet('mysheet1')
wb.create_sheet('mysheet2')
# 插入到0的位置
wb.create_sheet('mysheet3', 0)

print(wb.active) # <Worksheet "mysheet3">
print('======type', type(wb.active)) # <class 'openpyxl.worksheet.worksheet.Worksheet'>

print(wb.sheetnames) # ['mysheet3', 'Sheet', 'mysheet1', 'mysheet2']
print(wb.active.title) # mysheet3
print(wb['Sheet']) # <Worksheet "Sheet">

print('=============删除')
wb.remove(wb['Sheet'])
del wb['mysheet1']

print('===============重命名')
wb['mysheet2'].title = 'mysheet2_rename'

print('=============更改sheet tab颜色')
wb['mysheet2_rename'].sheet_properties.tabColor = '0077AA'

wb.save('test03_sheet.xlsx')
