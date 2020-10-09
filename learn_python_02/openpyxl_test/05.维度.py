from openpyxl import Workbook

wb = Workbook()
sheet = wb.active

sheet['A3'] = 39
sheet['B3'] = 19

rows = [
    (88, 46),
    (89, 38),
    (23, 59),
    (56, 21),
    (24, 18),
    (34, 15)
]

for row in rows:
    sheet.append(row)

print(sheet.dimensions) # A3:B9
print(type(sheet.dimensions)) # str
print(sheet.min_row, sheet.max_row) # 3 9
print(sheet.min_column, sheet.max_column) # 1 2

print('===========利用维度遍历')
for a, b in sheet[sheet.dimensions]:
	print(a.row, a.value, b.value)

	

