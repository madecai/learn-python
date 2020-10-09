from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
sheet = wb.active

# 合并单元格，只会保留左上角的单元格内容
sheet.merge_cells('A1:B2')
# ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=4)
# ws.unmerge_cells(start_row=2,start_column=1,end_row=2,end_column=4)

cell = sheet.cell(row=1, column=1)
cell.value = 'xxx'
# 水平居中、垂直居中
cell.alignment = Alignment(horizontal='center', vertical='center')

wb.save('test04_合并单元格.xlsx')